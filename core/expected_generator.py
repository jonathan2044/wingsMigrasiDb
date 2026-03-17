# Copyright (c) 2026 Jonathan Narendra - PT Naraya Prisma Digital
# Website : https://narayadigital.co.id
# All rights reserved.
"""
core/expected_generator.py

Engine untuk menghasilkan file EKSPEKTASI MIGRASI dari data sumber (kiri).

Menerapkan seluruh normalisasi + column transform rules + group-expansion
yang sama persis dengan CompareEngine, kemudian menghasilkan file output
dengan nama kolom TARGET (right_col), sehingga user dapat:
  - Langsung membandingkan ekspektasi vs data aktual di sistem target
    menggunakan VLOOKUP atau pivot
  - Melihat seharusnya data setelah migrasi tampak seperti apa

Tiga mode yang didukung:
  Standard / Column Expansion : 1 baris sumber → 1 baris output
  Row Expansion (GE 1:N)      : 1 baris sumber → N baris output sesuai mapping

Performa:
  CSV   — Python csv.writer + fetchmany(10_000).  utf-8-sig BOM agar langsung
           dibuka benar di Excel.  Tidak ada batasan baris.
  Excel — openpyxl write_only + fetchmany(10_000). Tidak ada Python row-loop
           besar.  Otomatis berhenti di EXCEL_ROW_LIMIT baris.
"""

from __future__ import annotations

import csv
import logging
import os
import tempfile
import uuid
from pathlib import Path
from typing import Callable, List, Optional, TYPE_CHECKING

import duckdb

from models.compare_config import (
    CompareConfig,
    ColumnTransformRule,
    GroupExpansionRule,
)
from core.normalization_engine import NormalizationEngine

if TYPE_CHECKING:
    from config.settings import AppSettings

logger = logging.getLogger(__name__)

ProgressCallback = Callable[[str, int, int], None]
CancelCallback   = Callable[[], bool]

EXCEL_ROW_LIMIT = 1_048_575   # maks baris di Excel (tidak termasuk header)
EXPORT_BATCH    = 10_000      # baris per fetchmany


class ExpectedMigrationGenerator:
    """
    Bangun file ekspektasi migrasi dari data sumber (kiri).

    Alur kerja:
      1. Import sumber kiri ke koneksi DuckDB sementara.
      2. Buat view normalized_left (normalisasi + transform rules).
      3. Jika GE rule aktif → bangun tabel mapping temp _ge_expected.
      4. Bangun SQL output yang me-rename kolom ke nama TARGET (right_col).
      5. Export ke CSV (csv.writer) atau Excel (openpyxl write_only).
    """

    def __init__(
        self,
        config: CompareConfig,
        transform_rules: Optional[List[ColumnTransformRule]] = None,
        group_expansion_rules: Optional[List[GroupExpansionRule]] = None,
        progress_cb: Optional[ProgressCallback] = None,
        cancel_cb: Optional[CancelCallback] = None,
    ):
        self._config   = config
        self._tx_rules = transform_rules or []
        self._ge_rules = group_expansion_rules or []
        self._emit     = progress_cb or (lambda *_: None)
        self._cancel   = cancel_cb   or (lambda: False)
        self._conn: Optional[duckdb.DuckDBPyConnection] = None
        self._norm     = NormalizationEngine(config.options)

    # ──────────────────────────────────────────────────────────────── public

    def generate(
        self,
        output_path: str,
        fmt: str,
        settings: "AppSettings",
    ) -> int:
        """
        Jalankan proses generasi ekspektasi.

        Args:
            output_path : path lengkap file output
            fmt         : 'csv' atau 'xlsx'
            settings    : AppSettings (untuk baca sumber + chunk size)

        Returns:
            Jumlah baris yang berhasil ditulis ke output.

        Raises:
            InterruptedError : jika cancel_cb() mengembalikan True
            Exception        : jika ada kesalahan selama generate
        """
        tmp_db = os.path.join(
            tempfile.gettempdir(),
            f"sfa_exp_{uuid.uuid4().hex}.duckdb",
        )
        try:
            self._conn = duckdb.connect(tmp_db)
            self._tune_conn()

            self._emit("Mengimpor data sumber...", 0, 0)
            left_rows = self._import_left(settings)
            logger.info("[expected] Import selesai: %d baris", left_rows)

            self._emit("Menerapkan transformasi kolom...", 0, left_rows)
            self._build_normalized_view()

            active_ge = self._find_active_ge_rule()
            if active_ge:
                self._emit("Membangun tabel ekspansi grup...", 0, left_rows)
                self._build_ge_table(active_ge)
                logger.info(
                    "[expected] GE aktif: %s → %s  (%d nilai kiri)",
                    active_ge.left_col, active_ge.right_cols, len(active_ge.mapping),
                )

            fmt_label = "CSV" if fmt == "csv" else "Excel"
            self._emit(f"Mengekspor ke {fmt_label}...", 0, left_rows)

            out_sql = self._build_output_sql(active_ge)
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)

            if fmt == "csv":
                n = self._export_csv(out_sql, output_path, left_rows)
            else:
                n = self._export_excel(out_sql, output_path, left_rows)

            self._emit("Selesai.", n, n)
            logger.info("[expected] Export selesai: %d baris → %s", n, output_path)
            return n

        finally:
            if self._conn:
                try:
                    self._conn.close()
                except Exception:
                    pass
                self._conn = None
            try:
                os.unlink(tmp_db)
            except OSError:
                pass

    # ──────────────────────────────────────────────────────────────── setup

    def _tune_conn(self):
        import os as _os
        cpu   = max(2, _os.cpu_count() or 4)
        self._conn.execute(f"PRAGMA threads={cpu}")
        try:
            import psutil
            avail = psutil.virtual_memory().available // (1024 * 1024)
        except Exception:
            avail = 4_096
        limit = max(2_048, min(int(avail * 0.60), 12 * 1_024))
        self._conn.execute(f"PRAGMA memory_limit='{limit}MB'")

    # ──────────────────────────────────────────────────────────────── import

    def _import_left(self, settings: "AppSettings") -> int:
        src   = self._config.left_source
        chunk = getattr(settings, "import_chunk_size", 50_000)

        def _cb(rows: int):
            self._emit(f"Mengimpor data sumber: {rows:,} baris", rows, 0)

        if src.source_type == "excel":
            from services.file_reader import ExcelReader
            return ExcelReader(src.file_path).import_to_duckdb(
                self._conn, "src_left",
                sheet_name=src.sheet_name or 0,
                skip_rows=src.skip_rows,
                chunk_size=chunk,
                progress_callback=_cb,
            )

        if src.source_type == "csv":
            from services.file_reader import CSVReader
            return CSVReader(
                src.file_path,
                separator=src.csv_separator,
                encoding=src.csv_encoding,
            ).import_to_duckdb(
                self._conn, "src_left",
                chunk_size=chunk,
                progress_callback=_cb,
            )

        if src.source_type in ("postgres", "mysql"):
            profile = self._resolve_profile(src, settings)
            if src.source_type == "mysql":
                from services.mysql_connector import MySQLConnector
                connector = MySQLConnector.from_profile(profile)
            else:
                from services.postgres_connector import PostgresConnector
                connector = PostgresConnector.from_profile(profile)
            try:
                return connector.import_to_duckdb(
                    self._conn, "src_left",
                    schema=src.schema_name,
                    table=src.table_name,
                    custom_query=src.custom_query if src.use_custom_query else "",
                    chunk_size=chunk,
                    progress_callback=_cb,
                )
            finally:
                connector.close()

        raise ValueError(f"Tipe sumber tidak didukung: {src.source_type!r}")

    def _resolve_profile(self, src, settings: "AppSettings"):
        from models.connection_profile import ConnectionProfile
        if src.connection_id:
            from storage.connection_store import ConnectionStore
            from storage.duckdb_storage import DuckDBStorage
            cs = ConnectionStore(DuckDBStorage(settings.db_path))
            p  = cs.get_by_id(src.connection_id)
            if not p:
                raise ValueError(f"Profil koneksi tidak ditemukan: {src.connection_id}")
            return p
        if getattr(src, "pg_connection_inline", None):
            return ConnectionProfile.from_dict(src.pg_connection_inline)
        raise ValueError("Tidak ada informasi koneksi database.")

    # ──────────────────────────────────────────────────────────────── normalized view

    def _build_normalized_view(self):
        """
        Buat view normalized_left dari src_left.
        Sama dengan CompareEngine._build_normalized_views() tapi hanya sisi kiri.
        """
        key_cols = [m.left_col for m in self._config.key_columns]
        val_cols = list(dict.fromkeys(m.left_col for m in self._config.compare_columns))

        # Pastikan GE left_col ada di val_cols agar bisa di-JOIN nanti
        ge = self._find_active_ge_rule()
        if ge and ge.left_col not in val_cols and ge.left_col not in key_cols:
            val_cols.append(ge.left_col)

        parts = []
        for col in key_cols:
            parts.append(
                f'TRIM(CAST("src_left"."{col}" AS VARCHAR)) AS "key_{col}"'
            )
        for col in val_cols:
            rules = [
                r for r in self._tx_rules
                if r.enabled
                and r.column_name.lower() == col.lower()
                and r.side in ("left", "both")
            ]
            expr = self._norm._build_expr_for_table_col("src_left", col, rules)
            parts.append(f'{expr} AS "left_{col}"')
        parts.append("ROW_NUMBER() OVER () AS left_rownum")

        self._conn.execute("DROP VIEW IF EXISTS normalized_left")
        self._conn.execute(
            "CREATE VIEW normalized_left AS SELECT "
            + ", ".join(parts)
            + " FROM src_left"
        )

    # ──────────────────────────────────────────────────────────────── GE rule

    def _find_active_ge_rule(self) -> Optional[GroupExpansionRule]:
        if not self._ge_rules:
            return None
        cmp_left  = {cm.left_col.lower() for cm in self._config.compare_columns}
        key_lower = {m.left_col.lower()  for m in self._config.key_columns}
        for rule in self._ge_rules:
            if not rule.enabled or not rule.mapping or not rule.right_cols:
                continue
            lc = rule.left_col.lower()
            if lc in key_lower:
                continue
            if lc in cmp_left:
                return rule
        return None

    def _build_ge_table(self, rule: GroupExpansionRule):
        """Isi tabel temp _ge_expected dengan semua baris mapping ekspansi."""
        n = len(rule.right_cols)
        self._conn.execute("DROP TABLE IF EXISTS _ge_expected")
        col_defs = ", ".join(f'"rc_{i}" VARCHAR' for i in range(n))
        self._conn.execute(
            f"CREATE TEMP TABLE _ge_expected (left_val VARCHAR, {col_defs})"
        )
        rows = []
        for lv, row_list in rule.mapping.items():
            for rv in row_list:
                padded = list(rv)[:n] + [""] * max(0, n - len(rv))
                rows.append([str(lv)] + [str(x) for x in padded])
        if rows:
            ph = ", ".join(["?"] * (n + 1))
            self._conn.executemany(
                f"INSERT INTO _ge_expected VALUES ({ph})", rows
            )
        cnt = self._conn.execute("SELECT COUNT(*) FROM _ge_expected").fetchone()[0]
        logger.info(
            "[expected] _ge_expected: %d baris dari %d nilai kiri",
            cnt, len(rule.mapping),
        )

    # ──────────────────────────────────────────────────────────────── output SQL

    def _build_output_sql(
        self, active_ge: Optional[GroupExpansionRule]
    ) -> str:
        """
        Bangun SELECT SQL untuk output ekspektasi.
        Kolom output menggunakan nama kolom TARGET (right_col dari mapping).
        """
        km = self._config.key_columns
        cm = self._config.compare_columns
        if active_ge:
            return self._sql_ge(active_ge, km, cm)
        return self._sql_standard(km, cm)

    def _sql_standard(self, km, cm) -> str:
        """Standard / Column Expansion: 1 row in → 1 row out."""
        parts = []
        for m in km:
            parts.append(f'"key_{m.left_col}" AS "{m.right_col}"')
        for m in cm:
            parts.append(f'"left_{m.left_col}" AS "{m.right_col}"')
        return (
            "SELECT " + ", ".join(parts)
            + " FROM normalized_left ORDER BY left_rownum"
        )

    def _sql_ge(self, rule: GroupExpansionRule, km, cm) -> str:
        """
        GE / Row Expansion: 1 row in → N rows out via mapping.

        Cabang A: baris yang left_val ADA di mapping → expand 1:N,
                  GE right_cols diambil dari _ge_expected.
        Cabang B: baris yang left_val TIDAK ada di mapping → fallback 1:1,
                  right_col[0] = nilai asli kiri, right_col[1..N-1] = NULL.
        """
        ge_lc      = rule.left_col
        ge_lc_low  = ge_lc.lower()
        n          = len(rule.right_cols)
        non_ge     = [m for m in cm if m.left_col.lower() != ge_lc_low]

        key_parts = [f'nl."key_{m.left_col}" AS "{m.right_col}"' for m in km]
        cmp_parts = [
            f'nl."left_{m.left_col}" AS "{m.right_col}"' for m in non_ge
        ]

        # Cabang A: expanded rows (INNER JOIN ke mapping)
        ge_a = [f'ge."rc_{i}" AS "{rule.right_cols[i]}"' for i in range(n)]
        sql_a = (
            "SELECT " + ", ".join(key_parts + cmp_parts + ge_a)
            + f' FROM normalized_left nl'
            f' INNER JOIN _ge_expected ge'
            f' ON nl."left_{ge_lc}" = ge.left_val'
        )

        # Cabang B: fallback rows (tidak ada di mapping)
        ge_b = [f'nl."left_{ge_lc}" AS "{rule.right_cols[0]}"']
        ge_b += [f'NULL AS "{rule.right_cols[i]}"' for i in range(1, n)]
        sql_b = (
            "SELECT " + ", ".join(key_parts + cmp_parts + ge_b)
            + f' FROM normalized_left nl'
            f' LEFT JOIN _ge_expected ge'
            f' ON nl."left_{ge_lc}" = ge.left_val'
            f' WHERE ge.left_val IS NULL'
        )

        return f"({sql_a}) UNION ALL ({sql_b})"

    # ──────────────────────────────────────────────────────────────── export

    def _export_csv(
        self, out_sql: str, output_path: str, estimated_rows: int
    ) -> int:
        """
        Export via Python csv.writer + utf-8-sig BOM.
        BOM diperlukan agar Microsoft Excel langsung mengenali encoding.
        Progress dilaporkan tiap EXPORT_BATCH baris.
        """
        cursor  = self._conn.execute(out_sql)
        headers = [d[0] for d in cursor.description]

        with open(output_path, "w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.writer(fh)
            writer.writerow(headers)
            total = 0
            while True:
                if self._cancel():
                    raise InterruptedError("Export dibatalkan oleh user.")
                batch = cursor.fetchmany(EXPORT_BATCH)
                if not batch:
                    break
                writer.writerows(batch)
                total += len(batch)
                self._emit(
                    f"Mengekspor CSV: {total:,} baris...",
                    total,
                    max(estimated_rows, total),
                )
        return total

    def _export_excel(
        self, out_sql: str, output_path: str, estimated_rows: int
    ) -> int:
        """
        Export Excel via openpyxl write_only mode.
        Berhenti otomatis saat total baris mencapai EXCEL_ROW_LIMIT.
        """
        try:
            import openpyxl
            from openpyxl.cell import WriteOnlyCell
            from openpyxl.styles import PatternFill, Font, Alignment
        except ImportError:
            raise RuntimeError(
                "Library openpyxl dibutuhkan untuk export Excel. "
                "Jalankan: pip install openpyxl"
            )

        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet("Ekspektasi Migrasi")

        cursor  = self._conn.execute(out_sql)
        headers = [d[0] for d in cursor.description]

        # ── Header row dengan styling ──
        hdr_fill  = PatternFill(fill_type="solid", fgColor="1E3A5F")
        hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        hdr_align = Alignment(horizontal="center", vertical="center")
        hdr_cells = []
        for h in headers:
            c = WriteOnlyCell(ws, value=h)
            c.fill      = hdr_fill
            c.font      = hdr_font
            c.alignment = hdr_align
            hdr_cells.append(c)
        ws.append(hdr_cells)

        # ── Data rows ──
        total     = 0
        truncated = False
        while True:
            if self._cancel():
                raise InterruptedError("Export dibatalkan oleh user.")
            batch = cursor.fetchmany(EXPORT_BATCH)
            if not batch:
                break
            for row in batch:
                ws.append([v if v is not None else "" for v in row])
                total += 1
                if total >= EXCEL_ROW_LIMIT:
                    truncated = True
                    break
            if truncated:
                break
            self._emit(
                f"Mengekspor Excel: {total:,} baris...",
                total,
                max(estimated_rows, total),
            )

        wb.save(output_path)

        if truncated:
            logger.warning(
                "[expected] Batas Excel (%d baris) tercapai — %d baris ter-ekspor.",
                EXCEL_ROW_LIMIT, total,
            )
        return total
