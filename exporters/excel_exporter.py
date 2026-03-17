# Copyright (c) 2026 Jonathan Narendra - PT Naraya Prisma Digital
# Website : https://narayadigital.co.id
"""
exporters/excel_exporter.py
Export hasil komparasi ke Excel — highlight warna per status, support multi-sheet.
"""

from __future__ import annotations
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional

from config.constants import (
    RESULT_MATCH, RESULT_MISMATCH,
    RESULT_MISSING_LEFT, RESULT_MISSING_RIGHT, RESULT_DUPLICATE_KEY,
    RESULT_STATUS_LABELS,
)

logger = logging.getLogger(__name__)

# Warna latar per status (hex ARGB untuk openpyxl)
_STATUS_FILL = {
    RESULT_MATCH:         "FFD1FAE5",   # hijau muda
    RESULT_MISMATCH:      "FFFEE2E2",   # merah muda
    RESULT_MISSING_LEFT:  "FFFEF3C7",   # kuning muda
    RESULT_MISSING_RIGHT: "FFFCE7F3",   # pink muda
    RESULT_DUPLICATE_KEY: "FFFFF7ED",   # oranye muda
}


class ExcelExporter:
    """Export hasil perbandingan ke file Excel (.xlsx)."""

    def __init__(self, output_path: str):
        self._path = Path(output_path)

    def export(
        self,
        records: List[Dict[str, Any]],
        summary: Optional[Dict[str, Any]] = None,
        job_name: str = "",
    ) -> str:
        """
        Export records ke Excel.
        Returns path file yang dihasilkan.
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("Butuh library openpyxl. Pastikan sudah terinstall.")

        wb = openpyxl.Workbook()

        # ---- Sheet 1: Ringkasan ----
        ws_summary = wb.active
        ws_summary.title = "Ringkasan"
        self._write_summary_sheet(ws_summary, summary, job_name)

        # ---- Sheet 2: Detail Hasil ----
        ws_detail = wb.create_sheet("Detail Hasil")
        self._write_detail_sheet(ws_detail, records)

        # Simpan
        self._path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(self._path))
        logger.info("Export Excel selesai: %s", self._path)
        return str(self._path)

    # ------------------------------------------------------------------ sheet writers

    def _write_summary_sheet(self, ws, summary: Optional[Dict], job_name: str):
        from openpyxl.styles import Font, PatternFill, Alignment

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=11)

        ws["A1"] = "SFA Compare Tool - Ringkasan Hasil"
        ws["A1"].font = title_font
        ws["A2"] = f"Nama Job: {job_name}"
        ws["A2"].font = Font(size=10)

        if summary:
            ws["A4"] = "Status"
            ws["B4"] = "Jumlah Baris"
            ws["A4"].font = header_font
            ws["B4"].font = header_font

            row = 5
            status_order = [
                RESULT_MATCH, RESULT_MISMATCH,
                RESULT_MISSING_LEFT, RESULT_MISSING_RIGHT, RESULT_DUPLICATE_KEY,
            ]
            total = summary.get("total_rows", 0)
            for s in status_order:
                count = summary.get(s, 0)
                pct = f"{count/total*100:.1f}%" if total else "0%"
                ws.cell(row=row, column=1, value=RESULT_STATUS_LABELS.get(s, s))
                ws.cell(row=row, column=2, value=count)
                ws.cell(row=row, column=3, value=pct)

                fill = PatternFill("solid", fgColor=_STATUS_FILL.get(s, "FFFFFFFF")[2:])
                ws.cell(row=row, column=1).fill = fill
                row += 1

            ws.cell(row=row + 1, column=1, value="Total Baris").font = header_font
            ws.cell(row=row + 1, column=2, value=total).font = header_font

    def _write_detail_sheet(self, ws, records: List[Dict[str, Any]]):
        from openpyxl.styles import Font, PatternFill, Alignment

        if not records:
            ws["A1"] = "Tidak ada data hasil untuk ditampilkan."
            return

        # Header
        headers = ["No", "Status", "Key", "Data Kiri", "Data Kanan", "Kolom Berbeda"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="1e3a5f"[0:6] if False else "1E3A5F")

        # Lebar kolom
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 45
        ws.column_dimensions["E"].width = 45
        ws.column_dimensions["F"].width = 35

        for row_idx, rec in enumerate(records, 2):
            status = rec.get("status", "")
            fill_color = _STATUS_FILL.get(status, "FFFFFFFF")[2:]  # strip FF prefix
            fill = PatternFill("solid", fgColor=fill_color)

            import json
            key_str = ", ".join(
                f"{k}={v}" for k, v in (rec.get("key_values") or {}).items()
            )
            left_str = json.dumps(rec.get("left_data") or {}, ensure_ascii=False)
            right_str = json.dumps(rec.get("right_data") or {}, ensure_ascii=False)
            diff_str = ", ".join(rec.get("diff_columns") or [])

            row_data = [
                row_idx - 1,
                RESULT_STATUS_LABELS.get(status, status),
                key_str,
                left_str,
                right_str,
                diff_str,
            ]

            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")
